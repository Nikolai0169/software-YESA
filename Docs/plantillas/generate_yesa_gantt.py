from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from pathlib import Path
from datetime import datetime, timedelta

out_path = Path(r"c:\Users\joser\OneDrive\Documentos\GitHub\software-YESA\Docs\plantillas\YESA_Gantt.xlsx")
out_path.parent.mkdir(parents=True, exist_ok=True)

wb = Workbook()
ws = wb.active
ws.title = 'Resumen YESA'

ws['A1'] = 'Proyecto'
ws['B1'] = 'YESA - E-commerce personalizado'
ws['A2'] = 'Objetivo'
ws['B2'] = 'Desarrollar una plataforma de comercio electrónico especializada en productos personalizados para orfebrería y cerámica artesanal.'
ws['A3'] = 'Justificación'
ws['B3'] = 'Resolver la falta de experiencias de compra personalizables y la necesidad de productos únicos, con gestión integral de catálogo, pedidos y administración.'
ws['A4'] = 'Feedback clave'
ws['B4'] = 'Priorizar personalización, carrito de compras, panel administrativo y experiencia de usuario fluida para clientes y equipo interno.'
ws['A5'] = 'Metodología'
ws['B5'] = 'Desarrollo iterativo con enfoque en backend, frontend, pruebas, documentación y puesta en marcha.'

for row in range(1, 6):
    for col in ['A', 'B']:
        ws[f'{col}{row}'].font = Font(bold=(col == 'A'), size=11)
        ws[f'{col}{row}'].alignment = Alignment(vertical='top', wrap_text=True)

ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 95


def render_progress(percent):
    blocks = 10
    filled = int(round(percent / 100 * blocks))
    filled = max(0, min(blocks, filled))
    return '█' * filled + '░' * (blocks - filled)


def badge_for_status(status):
    if status == 'Completado':
        return PatternFill('solid', fgColor='C6EFCE')
    if status == 'En curso':
        return PatternFill('solid', fgColor='FFE699')
    return PatternFill('solid', fgColor='F4CCCC')


def parse_date(s):
    try:
        return datetime.strptime(s, '%Y-%m-%d').date()
    except Exception:
        return None


def parse_weeks(s):
    if not s or not isinstance(s, str):
        return None
    s = s.lower()
    if 'semana' in s:
        try:
            return int(s.split()[0])
        except Exception:
            return None
    return None


ws2 = wb.create_sheet('Diagrama Gantt')
base_headers = [
    'Actividad',
    'Trimestre',
    'Inicio planeado',
    'Duración planeada/esperada',
    'Finalización planeada/esperada',
    'Inicio real',
    'Duración real',
    'Finalización real',
    '% completado',
    'Responsables',
    'Estado',
    'Barra visual',
]

# Rows: (Actividad, Trimestre, inicio_planeado, dur_plan, fin_plan, inicio_real, dur_real, fin_real, pct, responsables, estado, barra)
rows = [
    ['Nombre proyecto - Objetivo general - Objetivos específicos - Planteamiento del problema y pregunta problema - Alcance del proyecto - Justificación', 'T1 2025', '2025-04-29', '2 semanas', '2025-05-13', '2025-04-29', '2 semanas', '2025-05-13', 100, 'Ian — Líder / analista', 'Completado', render_progress(100)],
    ['Mapa de procesos BPMN del negocio', 'T1 2025', '2025-05-01', '2 semanas', '2025-05-15', '2025-05-01', '2 semanas', '2025-05-15', 100, 'Daniel — Analista funcional', 'Completado', render_progress(100)],
    ['Técnicas de recolección de información y análisis estadístico', 'T1 2025', '2025-05-15', '2 semanas', '2025-05-29', '2025-05-15', '2 semanas', '2025-05-29', 100, 'Ian — Líder / analista, Daniel — Analista funcional', 'Completado', render_progress(100)],
    ['Requerimientos funcionales y no funcionales (IEEE 830/1233/29148 o historias de usuario)', 'T1 2025', '2025-05-20', '3 semanas', '2025-06-10', '2025-05-20', '3 semanas', '2025-06-10', 100, 'Ian — Líder / analista, Daniel — Analista funcional', 'Completado', render_progress(100)],
    ['Validación de requerimientos con prototipo/mockups/wireframes', 'T1 2025', '2025-06-01', '2 semanas', '2025-06-15', '2025-06-01', '2 semanas', '2025-06-15', 100, 'Sebastián — Diseñador UX / frontend, Ian — Líder / analista', 'Completado', render_progress(100)],
    ['Uso de sistemas de control de versiones', 'T1 2025', '2025-04-29', '6 semanas', '2025-06-10', '2025-04-29', '6 semanas', '2025-06-10', 100, 'Raúl — Desarrollador backend, Sebastián — Desarrollador frontend', 'Completado', render_progress(100)],
    ['Formatos de fichas técnicas y estimación de costos', 'T2 2025', '2025-06-16', '2 semanas', '2025-06-30', '2025-06-16', '2 semanas', '2025-06-30', 100, 'Daniel — Analista funcional, Ian — Líder / analista', 'Completado', render_progress(100)],
    ['Diagrama de casos de uso y documentación extendida', 'T2 2025', '2025-07-01', '2 semanas', '2025-07-15', '2025-07-01', '2 semanas', '2025-07-15', 100, 'Ian — Líder / analista, Daniel — Analista funcional', 'Completado', render_progress(100)],
    ['Modelo entidad relación con notación crow’s foot', 'T2 2025', '2025-07-01', '2 semanas', '2025-07-15', '2025-07-01', '2 semanas', '2025-07-15', 100, 'Raúl — Desarrollador backend', 'Completado', render_progress(100)],
    ['Normalización del modelo relacional (3FN)', 'T2 2025', '2025-07-16', '2 semanas', '2025-07-30', '2025-07-16', '2 semanas', '2025-07-30', 100, 'Raúl — Desarrollador backend', 'Completado', render_progress(100)],
    ['Diccionario de datos', 'T2 2025', '2025-07-16', '2 semanas', '2025-07-30', '2025-07-16', '2 semanas', '2025-07-30', 100, 'Raúl — Desarrollador backend, Daniel — Analista funcional', 'Completado', render_progress(100)],
    ['Diagrama de clases UML 2.4.1', 'T2 2025', '2025-08-01', '2 semanas', '2025-08-15', '2025-08-01', '2 semanas', '2025-08-15', 100, 'Raúl — Desarrollador backend, Sebastián — Desarrollador frontend', 'Completado', render_progress(100)],
    ['Diagrama de despliegue UML 2.4.1', 'T2 2025', '2025-08-01', '2 semanas', '2025-08-15', '2025-08-01', '2 semanas', '2025-08-15', 100, 'Raúl — Desarrollador backend', 'Completado', render_progress(100)],
    ['Uso de sistemas de control de versiones', 'T2 2025', '2025-06-16', '8 semanas', '2025-08-15', '2025-06-16', '8 semanas', '2025-08-15', 100, 'Raúl — Desarrollador backend, Sebastián — Desarrollador frontend', 'Completado', render_progress(100)],
    ['Construcción de la base de datos con DDL/Schema Validation', 'T3 2025', '2025-08-16', '4 semanas', '2025-09-13', '2025-08-16', '4 semanas', '2025-09-13', 100, 'Raúl — Desarrollador backend', 'Completado', render_progress(100)],
    ['Uso de la base de datos con DML/CRUD y consultas', 'T3 2025', '2025-09-01', '3 semanas', '2025-09-20', '2025-09-01', '3 semanas', '2025-09-20', 100, 'Raúl — Desarrollador backend', 'Completado', render_progress(100)],
    ['Seguridad de la base de datos: encriptación y protección de datos', 'T3 2025', '2025-09-10', '2 semanas', '2025-09-24', '2025-09-10', '2 semanas', '2025-09-24', 100, 'Raúl — Desarrollador backend', 'Completado', render_progress(100)],
    ['Front-end funcional con framework web y autenticación JWT', 'T3 2025', '2025-09-15', '4 semanas', '2025-10-13', '2025-09-15', '4 semanas', '2025-10-13', 100, 'Sebastián — Desarrollador frontend, Ian — Líder / analista', 'Completado', render_progress(100)],
    ['Uso de sistemas de control de versiones', 'T3 2025', '2025-08-16', '8 semanas', '2025-10-13', '2025-08-16', '8 semanas', '2025-10-13', 100, 'Raúl — Desarrollador backend, Sebastián — Desarrollador frontend', 'Completado', render_progress(100)],
    ['Prototipo navegable del software (HTML/CSS)', 'T3 2025', '2025-09-20', '3 semanas', '2025-10-11', '2025-09-20', '3 semanas', '2025-10-11', 100, 'Sebastián — Desarrollador frontend', 'Completado', render_progress(100)],
    ['Implementación de la API REST agregando seguridad', 'T4 2025', '2025-10-14', '4 semanas', '2025-11-11', '2025-10-14', '4 semanas', '2025-11-11', 100, 'Raúl — Desarrollador backend', 'Completado', render_progress(100)],
    ['Front-end web consumiendo la API REST (avance 80%)', 'T4 2025', '2025-10-14', '4 semanas', '2025-11-11', '2025-10-14', '4 semanas', '2025-11-11', 100, 'Sebastián — Desarrollador frontend', 'Completado', render_progress(100)],
    ['Uso de sistemas de control de versiones', 'T4 2025', '2025-10-14', '6 semanas', '2025-11-25', '2025-10-14', '6 semanas', '2025-11-25', 100, 'Raúl — Desarrollador backend, Sebastián — Desarrollador frontend', 'Completado', render_progress(100)],
    ['Desarrollo de la codificación al 100%', 'T5 2025', '2025-11-26', '6 semanas', '2026-01-07', '2025-11-26', '6 semanas', '2026-01-07', 100, 'Raúl — Desarrollador backend, Sebastián — Desarrollador frontend', 'Completado', render_progress(100)],
    ['Consumo de la API REST con aplicaciones móviles', 'T5 2025', '2026-01-08', '3 semanas', '2026-01-29', '2026-01-08', '3 semanas', '2026-01-29', 100, 'Sebastián — Desarrollador frontend, Raúl — Desarrollador backend', 'Completado', render_progress(100)],
    ['Implementación de la API REST documentada (Swagger u otra herramienta)', 'T5 2025', '2026-01-08', '2 semanas', '2026-01-22', '2026-01-08', '2 semanas', '2026-01-22', 100, 'Raúl — Desarrollador backend, Daniel — Analista funcional', 'Completado', render_progress(100)],
    ['Aplicación de metodología ágil (historias de usuario, roles, sprints, backlog)', 'T5 2025', '2026-01-15', '3 semanas', '2026-02-05', '2026-01-15', '3 semanas', '2026-02-05', 100, 'Ian — Líder / analista, Daniel — Analista funcional', 'Completado', render_progress(100)],
    ['Técnicas de pruebas de software', 'T6 2025', '2026-06-01', '4 semanas', '2026-09-10', '2026-06-01', 'En curso', '', 35, 'Ian — Líder / analista, Raúl — Desarrollador backend, Sebastián — Desarrollador frontend', 'En curso', render_progress(35)],
    ['Plan de instalación, respaldo, migración y capacitación', 'T6 2025', '2026-06-15', '4 semanas', '2026-09-10', '2026-06-15', 'En curso', '', 30, 'Daniel — Analista funcional, Ian — Líder / analista, Sebastián — Desarrollador frontend', 'En curso', render_progress(30)],
    ['Modelo de calidad', 'T6 2025', '2026-06-20', '3 semanas', '2026-09-10', '2026-06-20', 'En curso', '', 25, 'Ian — Líder / analista, Daniel — Analista funcional', 'En curso', render_progress(25)],
    ['Manuales de instalación, técnico y de usuario', 'T6 2025', '2026-07-01', '4 semanas', '2026-09-10', '2026-07-01', 'En curso', '', 40, 'Daniel — Analista funcional, Sebastián — Desarrollador frontend', 'En curso', render_progress(40)],
    ['Despliegue del aplicativo según diseño de arquitectura UML', 'T6 2025', '2026-07-15', '4 semanas', '2026-09-10', '2026-07-15', 'En curso', '', 20, 'Raúl — Desarrollador backend, Sebastián — Desarrollador frontend', 'En curso', render_progress(20)],
    ['Uso de sistemas de control de versiones', 'T6 2025', '2026-06-01', '4 semanas', '2026-09-10', '2026-06-01', 'En curso', '', 45, 'Raúl — Desarrollador backend, Sebastián — Desarrollador frontend', 'En curso', render_progress(45)],
]

# compute calendar range (weekly columns)
planned_starts = [parse_date(r[2]) for r in rows if parse_date(r[2])]
planned_ends = [parse_date(r[4]) for r in rows if parse_date(r[4])]
start_date = min(planned_starts) if planned_starts else datetime.today().date()
end_date = max(planned_ends) if planned_ends else start_date + timedelta(weeks=52)

# ensure end_date at least the user's requested final (2026-09-10)
fallback_end = datetime.strptime('2026-09-10', '%Y-%m-%d').date()
if end_date < fallback_end:
    end_date = fallback_end

# build week starts
week_starts = []
cur = start_date
while cur <= end_date:
    week_starts.append(cur)
    cur = cur + timedelta(days=7)

headers = base_headers + [ws_date.strftime('%Y-%m-%d') for ws_date in week_starts]
ws2.append(headers)

thin = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)

# write rows and weekly fills
for row in rows:
    base = row[:]  # copy
    # append placeholder weekly cells, will fill later
    base = base + [''] * len(week_starts)
    ws2.append(base)

# style header
for cell in ws2[1]:
    cell.font = Font(bold=True, color='FFFFFF')
    cell.fill = PatternFill('solid', fgColor='1F4E78')
    cell.alignment = Alignment(horizontal='center', vertical='center')

# fill weekly bars
for r_idx in range(2, ws2.max_row + 1):
    act_start = parse_date(ws2.cell(row=r_idx, column=3).value)
    act_end = parse_date(ws2.cell(row=r_idx, column=5).value)
    dur_str = ws2.cell(row=r_idx, column=4).value
    weeks = parse_weeks(dur_str)
    status = ws2.cell(row=r_idx, column=11).value
    # fallback: if act_end missing but weeks present
    if not act_end and act_start and weeks:
        act_end = act_start + timedelta(weeks=weeks) - timedelta(days=1)
    if not act_start or not act_end:
        continue
    # color based on status
    if status == 'Completado':
        fill = PatternFill('solid', fgColor='9BE39B')
    elif status == 'En curso':
        fill = PatternFill('solid', fgColor='FFD966')
    else:
        fill = PatternFill('solid', fgColor='E7E6E6')

    # for each week column, fill if week interval intersects activity interval
    for w_i, w_start in enumerate(week_starts, start=1):
        w_col = len(base_headers) + w_i
        w_end = w_start + timedelta(days=6)
        if (w_start <= act_end) and (w_end >= act_start):
            ws2.cell(row=r_idx, column=w_col).fill = fill

# trim empty week columns (no activity fills)
first_week_col = len(base_headers) + 1
last_week_col = ws2.max_column
cols_to_delete = []
for col in range(first_week_col, last_week_col + 1):
    has_fill = False
    for r in range(2, ws2.max_row + 1):
        cell = ws2.cell(row=r, column=col)
        if getattr(cell.fill, 'fill_type', None) is not None:
            has_fill = True
            break
    if not has_fill:
        cols_to_delete.append(col)

# delete from right to left to keep indices valid
for col in reversed(cols_to_delete):
    ws2.delete_cols(col)

# apply borders and alignment for all cells
for row in ws2.iter_rows(min_row=1, max_row=ws2.max_row, max_col=ws2.max_column):
    for cell in row:
        cell.border = thin
        cell.alignment = Alignment(vertical='center', wrap_text=True)

# adjust week column widths (make them narrow) and freeze header row
new_last_col = ws2.max_column
for col_idx in range(len(base_headers) + 1, new_last_col + 1):
    letter = get_column_letter(col_idx)
    ws2.column_dimensions[letter].width = 3
ws2.freeze_panes = ws2['A2']

# rotate weekly header labels for readability
ws2.row_dimensions[1].height = 48
for col_idx in range(len(base_headers) + 1, ws2.max_column + 1):
    cell = ws2.cell(row=1, column=col_idx)
    cell.alignment = Alignment(textRotation=45, horizontal='center', vertical='bottom', wrap_text=True)

# add inline legend to the Diagrama Gantt sheet (right side)
legend_start = ws2.max_column + 2
ws2.cell(row=1, column=legend_start, value='Leyenda').font = Font(bold=True)
ws2.cell(row=2, column=legend_start, value='').fill = PatternFill('solid', fgColor='9BE39B')
ws2.cell(row=2, column=legend_start + 1, value='Completado')
ws2.cell(row=3, column=legend_start, value='').fill = PatternFill('solid', fgColor='FFD966')
ws2.cell(row=3, column=legend_start + 1, value='En curso')
ws2.cell(row=4, column=legend_start, value='').fill = PatternFill('solid', fgColor='E7E6E6')
ws2.cell(row=4, column=legend_start + 1, value='Sin asignar / otro')
ws2.column_dimensions[get_column_letter(legend_start)].width = 6
ws2.column_dimensions[get_column_letter(legend_start + 1)].width = 28

print('TRIMMED_WEEK_COLUMNS', len(cols_to_delete))

# keep separate legend sheet as reference
ws_legend = wb.create_sheet('Leyenda')
ws_legend['A1'] = 'Color'
ws_legend['B1'] = 'Significado'
ws_legend['A2'] = ''
ws_legend['B2'] = 'Completado'
ws_legend['A2'].fill = PatternFill('solid', fgColor='9BE39B')
ws_legend['A3'] = ''
ws_legend['B3'] = 'En curso'
ws_legend['A3'].fill = PatternFill('solid', fgColor='FFD966')
ws_legend['A4'] = ''
ws_legend['B4'] = 'Sin asignar / otro'
ws_legend['A4'].fill = PatternFill('solid', fgColor='E7E6E6')
for cell in ws_legend['A1:B1'][0]:
    cell.font = Font(bold=True)
ws_legend.column_dimensions['A'].width = 6
ws_legend.column_dimensions['B'].width = 36

# small visual sheet
ws3 = wb.create_sheet('Vista Gantt')
ws3['A1'] = 'Vista visual del Gantt - Proyecto YESA'
ws3['A1'].font = Font(bold=True, size=14)
headers_v = ['Actividad', 'Trimestre', 'Inicio planeado', 'Fin planeado', 'Inicio real', 'Fin real', 'Estado', 'Progreso']
for col, value in enumerate(headers_v, start=1):
    ws3.cell(row=3, column=col, value=value)
    ws3.cell(row=3, column=col).font = Font(bold=True)
    ws3.cell(row=3, column=col).fill = PatternFill('solid', fgColor='D9EAF7')

for i, row in enumerate(rows, start=4):
    ws3.cell(row=i, column=1, value=row[0])
    ws3.cell(row=i, column=2, value=row[1])
    ws3.cell(row=i, column=3, value=row[2])
    ws3.cell(row=i, column=4, value=row[4])
    ws3.cell(row=i, column=5, value=row[5])
    ws3.cell(row=i, column=6, value=row[7])
    ws3.cell(row=i, column=7, value=row[10])
    ws3.cell(row=i, column=8, value=row[11])

for row in ws3.iter_rows(min_row=3, max_row=ws3.max_row):
    for cell in row:
        cell.border = thin
        cell.alignment = Alignment(vertical='center', wrap_text=True)

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
    ws3.column_dimensions[col].width = 24 if col in {'A', 'H'} else 16

wb.save(out_path)
print(f'Created {out_path}')

# compute implementation duration (actividad que contiene "Implementación de la API REST agregando seguridad")
impl_days = None
for r in rows:
    if 'Implementación de la API REST agregando seguridad' in r[0]:
        s = parse_date(r[2])
        e = parse_date(r[4])
        if s and e:
            impl_days = (e - s).days
        else:
            weeks = parse_weeks(r[3])
            if weeks:
                impl_days = weeks * 7
        break

if impl_days is not None:
    print('IMPLEMENTACION_DIAS', impl_days)
